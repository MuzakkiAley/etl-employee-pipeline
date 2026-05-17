import pandas as pd
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'Muzzasql03#',  # ganti dengan password MySQL kamu
    'database': 'employee_db'
}

# ─── FETCH DATA ───────────────────────────────────────────
def fetch_data():
    print("[REPORT] Fetching data from MySQL...")
    conn = mysql.connector.connect(**DB_CONFIG)

    employees = pd.read_sql("SELECT * FROM employees", conn)

    dept_summary = pd.read_sql("""
        SELECT 
            department,
            COUNT(*) as total_employees,
            AVG(salary) as avg_salary,
            MIN(salary) as min_salary,
            MAX(salary) as max_salary,
            SUM(CASE WHEN status = 'Active' THEN 1 ELSE 0 END) as active_count
        FROM employees
        GROUP BY department
        ORDER BY total_employees DESC
    """, conn)

    conn.close()
    return employees, dept_summary

# ─── STYLE HELPERS ────────────────────────────────────────
def style_header(ws, row, col_count, color="1D6FA5"):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

# ─── GENERATE REPORT ──────────────────────────────────────
def generate_report(employees, dept_summary):
    filename = f"employee_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    print(f"[REPORT] Generating {filename}...")

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: All Employees
        employees.to_excel(writer, sheet_name='All Employees', index=False)

        # Sheet 2: Department Summary
        dept_summary['avg_salary'] = dept_summary['avg_salary'].round(0).astype(int)
        dept_summary.to_excel(writer, sheet_name='Department Summary', index=False)

        # Sheet 3: Active vs Inactive
        status_summary = employees.groupby('status').agg(
            total=('employee_id', 'count'),
            avg_salary=('salary', 'mean')
        ).reset_index()
        status_summary['avg_salary'] = status_summary['avg_salary'].round(0).astype(int)
        status_summary.to_excel(writer, sheet_name='Status Summary', index=False)

    # Apply styles
    wb = load_workbook(filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_count = ws.max_column
        style_header(ws, 1, col_count)
        auto_width(ws)

        # Zebra stripe rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="EEF4FA")

    wb.save(filename)
    print(f"  Report saved: {filename}")
    print(f"  Sheets: All Employees, Department Summary, Status Summary")
    return filename

# ─── MAIN ─────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 50)
    print("  REPORT GENERATOR — Employee Data")
    print("=" * 50)

    employees, dept_summary = fetch_data()
    filename = generate_report(employees, dept_summary)

    print(f"\n[DONE] Report generated successfully!")
    print(f"  Open '{filename}' to view the report.")